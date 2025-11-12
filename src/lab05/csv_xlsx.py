import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Преобразует CSV файл в формат XLSX.
    Используется библиотека openpyxl для работы с Excel.
    Первая строка исходного файла считается заголовком таблицы.
    Лист получает название "Sheet1".
    Для удобства просмотра устанавливается автоширина колонок с минимальным значением 8 символов.
    """
    # Преобразуем пути в объекты Path для удобной работы с файловой системой
    csv_file = Path(csv_path)
    excel_file = Path(xlsx_path)
    
    # Проверяем существование исходного CSV файла
    if not csv_file.exists():
        raise FileNotFoundError(f"Не удается найти CSV файл по указанному пути: {csv_file}")
    
    # Читаем данные из CSV файла
    csv_data = []
    try:
        # Открываем файл с автоматическим определением кодировки
        with csv_file.open('r', encoding='utf-8') as file:
            # Определяем разделитель по содержимому файла
            sample_content = file.read(1024)
            file.seek(0)  # Возвращаемся к началу файла
            dialect_detector = csv.Sniffer()
            csv_dialect = dialect_detector.sniff(sample_content)
            
            # Создаем reader с определенным разделителем
            csv_reader = csv.reader(file, dialect=csv_dialect)
            
            # Построчно считываем все данные
            for line in csv_reader:
                csv_data.append(line)
                
    except csv.Error as e:
        raise ValueError(f"Проблема с форматом CSV файла: {e}")
    except Exception as e:
        raise ValueError(f"Общая ошибка при обработке CSV: {e}")
    
    # Проверяем что файл не пустой
    if not csv_data:
        raise ValueError("CSV файл не содержит данных")
    
    # Создаем новую книгу Excel
    excel_workbook = Workbook()
    excel_sheet = excel_workbook.active
    excel_sheet.title = "Sheet1"  # Устанавливаем название листа
    
    # Переносим данные из CSV в Excel
    for row_data in csv_data:
        excel_sheet.append(row_data)
    
    # Настраиваем ширину колонок для лучшего отображения
    for column_index, column_cells in enumerate(excel_sheet.columns, 1):
        max_text_length = 0
        current_column_letter = get_column_letter(column_index)
        
        # Ищем самую длинную ячейку в колонке
        for cell in column_cells:
            try:
                if cell.value is not None:
                    # Обновляем максимальную длину если нашли большее значение
                    max_text_length = max(max_text_length, len(str(cell.value)))
            except:
                # Пропускаем ячейки с ошибками преобразования
                continue
        
        # Устанавливаем ширину колонки (минимум 8 символов)
        column_width = max(max_text_length + 2, 8)
        excel_sheet.column_dimensions[current_column_letter].width = column_width
    
    # Сохраняем результат
    excel_file.parent.mkdir(parents=True, exist_ok=True)  # Создаем папки если нужно
    excel_workbook.save(excel_file)